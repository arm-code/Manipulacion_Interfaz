import openpyxl
import pyautogui
import time

# ACTUALIZACION: NOVIEMBRE 2023
# EN EL ARCHIVO DE EXCEL, AHORA SE LEEN LOS DATOS MEDIANTE UNA FORMULA LA CUAL VINCULA LOS DEMAS ARCHIVOS DE LAS 26 PLANTILLAS
# PARA QUE FUNCIONE SE DEBE CAMBIAR EL NOMBRE DE DICHOS ARCHIVOS, PONERLE DEL 1 AL 26 COMO NOMBRE, PARA QUE LA FORMULA PUEDA 
# RECONOCERLOS FACILMENTE AL HACER UN RELLENO DE CELDA.

# MATERIAS PARA EL PLAN 22, VAN DESDE LA MATERIA 01 HASTA LA 26 (DE LA 21 SE SALTA A LA 26)
number_asignatura = ['01', '02','03','04','05','06','07','08','09','10','11','12','13','14','15','16','17','18','19','20','21','26']

# MATERIAS DEL PLAN 33, ESTAS DEPENDEN DE LAS MATERIAS QUE SE HAYAN SOLICITADO.
# HAY QUE TENER CUIDADO CON LAS MATERIAS QUE TIENEN MAS DE 30 PREGUNTAS. 
# ESTE PROGRAMA SOLO ESTA DISE;ADO PARA MATERIAS DE 30 PREGUNTAS!!!!
#number_asignatura = ['21', '34', '43', '53', '60']

# DEBEMOS CAMBIAR LA ETAPA, FASE Y EL PLAN
name_asignatura = '2312A22'

# AQUI SE CARGA EL ARCHIVO QUE CONTIENE TODAS LAS PLANTILLAS DE LAS DIFERENTES MATERIAS. 
excel = 'relleno22.xlsx'

workbook = openpyxl.load_workbook(excel, data_only=True)
sheet = workbook.active

num_cols = sheet.max_column
num_rows = sheet.max_row

print("COLOQUE EL CURSOR EN LA PANTALLA DE CAPTURA!")
time.sleep(5)
print('ESPERE MIENTRAS SE CARGAN LOS DATOS...')
n = 0

# # el siguiente bloque de codigo lee las celdas de manera horizontal (filas)
# for row_index in range(1, num_rows + 1):    
#     data_to_enter = []
#     for col_index in range(1, num_cols + 1):
#         cell_value = sheet.cell(row = row_index, column=col_index).value
#         #print(cell_value)
#         data_to_enter.append(str(cell_value))

# el sig bloque de codigo las lee de manera vertical (columnas)
for col_index in range(1, num_cols + 1):
    data_to_enter = []
    for row_index in range(1, num_rows + 1):
        cell_value = sheet.cell(row=row_index, column=col_index).value
        # print(cell_value)
        if cell_value and cell_value.startswith('='):
            cell_value = sheet.cell(row=row_index, column=col_index).value
        data_to_enter.append(str(cell_value))
    # Hacer algo con los datos de la columna actual, por ejemplo, imprimirlos
    print(data_to_enter)
    
    #se deben verificar las coordenadas del click, de lo contrario se va ir a otro lado la captura
    pyautogui.click(x=100, y= 80)

    # NUMERO DE ASIGNATURA
    pyautogui.write(number_asignatura[n])
    pyautogui.press('tab')

    # CLAVE DE PLANTILLA
    pyautogui.write(number_asignatura[n] + name_asignatura)
    pyautogui.press('tab')

    # NUMERO DE RESPUESTAS
    pyautogui.write('30')
    pyautogui.press('tab')

    #BASE DE CODIFICACION
    pyautogui.write('1')
    pyautogui.press('tab')
    
    # NORMAS DE CODIFICACION
    # LAS NORMAS DE CODIFICACION VARIAN PARA MATERIAS DE 30 Y 40 PREGUNTAS.
    # PARA MATERIAS DE 30 PREGUNTAS USAMOS [17,19,21,24,27]
    # PARA MATERIAS DE 40 PREGUNTAS USAMOS [23,27,31,34,37]

    pyautogui.write('17')   # NORMA 1
    pyautogui.press('tab')  
    pyautogui.write('19')   # NORMA 2
    pyautogui.press('tab')  
    pyautogui.write('21')   # NORMA 3
    pyautogui.press('tab')
    pyautogui.write('24')   # NORMA 4
    pyautogui.press('tab')
    pyautogui.write('27')   # NORMA 5
    pyautogui.press('tab')

    # RESPUESTAS
    for i in range(1,31):
        if i % 5 == 0 :
            pyautogui.write(data_to_enter[i-1])
            pyautogui.press('tab')
            
        else:
            pyautogui.write(data_to_enter[i-1])
    
    input('REVISE CUIDADOSAMENTE LA CAPTURA \nPRESIONE ENTER PARA CONTINUAR...')
    time.sleep(3)

    # GUARDAR PLANTILLA
    pyautogui.press('f2')
    pyautogui.press('enter')
    pyautogui.press('enter')
    print('CAPTURA EXITOSA!\n')  
    n = n + 1
    print('SIGUIENTE MATERIA = ', number_asignatura[n])
    print("\t****************")
    input('\tDESEA SALIR? \n\t-PRESIONE CTRL + C \n\tAGREGAR SIG MATERIA? \n\t-PRESIONE ENTER\n')
    time.sleep(5)
workbook.close()
