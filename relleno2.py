import time
import openpyxl
import pyautogui

PORTADA = """

 __________________________________
|              ____                |
|   ________  / / /__  ____  ____  |
|  / ___/ _ \/ / / _ \/ __ \/ __ \ |
| / /  /  __/ / /  __/ / / / /_/ / |
|/_/   \___/_/_/\___/_/ /_/\____/  |
|                                  |
|  by: arm-code (GPL License)      |
|__________________________________|
"""


print(PORTADA)

print('Bienvenido al programa para cargar las plantillas en el SIOSAD.')

option = 0
etapa = ''
fase = ''
plan = ''
name_asignatura = ''
excel = ''
normaCalificacion3340 = [23, 27, 31, 34, 37]  # para 40 preguntas del plan 33
# para plan 22 y 33 de 30 preguntas
normaCalificacion2233 = [17, 19, 21, 24, 27]
# Para plan 22 y 33 de 30 preguntas (2024)
normaCalificacion2233_2024 = [16, 18, 20, 23, 26]

while option != 1:
    etapa = input('Ingrese la etapa: ')
    fase = input('Ingrese la fase: ')
    plan = input('Ingrese el plan: ')
    name_asignatura = etapa + plan

    if plan == '33':
        print(
            'Las asignaturas de 40 preguntas tendra que meter manualmente las ultimas 10.')
        print('Aun no se implementa esa funcionalidad... :(')

    # DEBEMOS COLOCAR EL NOMBRE DE ARCHIVO QUE CONTIENE LAS PLANTILLAS
    name_file = input('Ingrese el nombre del archivo de excel: ')
    excel = name_file+'.xlsx'

    print('\n\t| Usted ha ingresado los sig. datos:')
    print(f'\t| Etapa:    { etapa } ')
    print(f'\t| Fase:     { fase }  ')
    print(f'\t| Plan:     { plan }  ')
    print(f'\t| Excel:    { excel } ')
    OPTION = int(input(
        '\t| Ingrese [0] y presione enter pa corregir los datos. \n\t| Ingrese [1] y presione enter para continuar         \n\t=> '))

# intentamos abrir el archivo con el nombre del excel proporcionado
print('Se esta leyendo el archivo de excel...')

try:
    workbook = openpyxl.load_workbook(excel, data_only=True)
    print("[Coloque el cursor en la ventana del SIOSAD]")
    print('[en es el espacio de la asignatura]')
    time.sleep(2)

    n = 0

    # Iterar sobre cada hoja del libro
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        num_cols = sheet.max_column
        num_rows = sheet.max_row

        print('\nLos siguientes datos han sido leidos del archivo:')
        print(f'Reactivos de la asignatura: #{sheet_name}')

        # Definir el rango de columnas desde B hasta AE
        columnas_a_leer = list(sheet.iter_cols(
            min_col=2, max_col=31, min_row=1, max_row=2))

        # Iterar sobre cada fila de la hoja
        for row_index in range(2, 3):  # Solo filas 1 y 2
            data_to_enter = []

            # Iterar sobre cada celda en el rango de columnas definido
            for col_index in range(len(columnas_a_leer)):
                cell_value = columnas_a_leer[col_index][row_index - 1].value
                data_to_enter.append(str(cell_value))

            # Imprimir los datos de la fila actual
            # print(data_to_enter)
            for i in range(len(data_to_enter)):
                if (i+1) % 5 == 0:
                    print('|', data_to_enter[i], '|')
                else:
                    print('|', data_to_enter[i], end=' ')
            print('\nINGRESANDO LOS DATOS EN EL SIOSAD...')
            time.sleep(2)
            # se deben verificar las coordenadas del click, de lo contrario se va ir a otro lado la captura
            pyautogui.click(x=150, y=150)

            # NUMERO DE ASIGNATURA
            # pyautogui.write(number_asignatura[n])
            pyautogui.write(sheet_name)
            pyautogui.press('tab')

            # CLAVE DE PLANTILLA
            pyautogui.write(sheet_name + name_asignatura)
            pyautogui.press('tab')

            # NUMERO DE RESPUESTAS
            pyautogui.write('30')
            pyautogui.press('tab')

            # BASE DE CODIFICACION
            pyautogui.write('1')
            pyautogui.press('tab')

            # NORMAS DE Calificacion
            try:
                for i in range(len(normaCalificacion2233_2024)):
                    pyautogui.write(str(normaCalificacion2233_2024[i]))
                    pyautogui.press('tab')

            except NameError:
                print('error al ingresar las normas')

            # Ingresando reactivos al SIOSAD
            for i in range(1, 31):
                if i % 5 == 0:
                    pyautogui.write(data_to_enter[i-1])
                    pyautogui.press('tab')

                else:
                    pyautogui.write(data_to_enter[i-1])

            input(
                '\nREVISE CUIDADOSAMENTE LA CAPTURA \nPRESIONE ENTER PARA CONTINUAR...\n>')
            print('enter')
            print(
                'EN SEGUIDA VUELVA A COLOCAR EL CURSOR EN LA VENTANA DEL SIOSAD PLANTILLAS.')
            time.sleep(2)

            # GUARDAR PLANTILLA
            pyautogui.press('f2')
            pyautogui.press('enter')
            pyautogui.press('enter')
            print('\nCAPTURA EXITOSA!\n')
            try:
                n = n + 1
                print('\t| Siguiente asignatura: ')
                input(
                    '\t| PARA DETENER EL PROGRAMA:   [PRESIONE CTRL + C]\n\t| PARA AGREGAR SIG MATERIA:   [PRESIONE ENTER] \n\t=>')

                print('enter')
                time.sleep(2)
            except TypeError:
                print('Al parecer se ha terminado de leer el EXCEL.')
        workbook.close()
except RuntimeError:
    print('Error al intentar abrir el archivo!!')
    print('Vuelva a ejecutar el programa!')
